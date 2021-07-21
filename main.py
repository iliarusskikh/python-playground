import random

import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    for item in range(2):
        randword = random.randint(1,sheet.max_row - 1 )
        cell = sheet.cell(randword,1)
        print(cell.value)
        print(input('Enter in the following format: präteretum perfect english '))
        output = ' '
        cell2 = sheet.cell(randword, 2)
        cell3 = sheet.cell(randword, 3)
        cell4 = sheet.cell(randword, 4)
        cell5 = sheet.cell(randword, 5)
        output = cell2.value + '_|_' + cell3.value + '_|_' + cell4.value + '_|_' + cell5.value
        print(output)


    # for row in range(2, sheet.max_row + 1):
    #     cell = sheet.cell(row,3)
    #     print(cell.value)
    #     corrected_price = cell.value * 0.9
    #     corrected_price_cell = sheet.cell(row,4)
    #     corrected_price_cell = corrected_price
    #
    # values = Reference(sheet, min_row=2, max_row= sheet.max_row, min_col=4,max_col=4)
    # chart = BarChart()
    # chart.add_data(values)
    # sheet.add_chart(chart, 'e2')
    #
    # wb.save(filename)



process_workbook('mygerman.xlsx')
